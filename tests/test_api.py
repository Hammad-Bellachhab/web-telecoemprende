import os
import unittest

os.environ["ADMIN_PASSWORD"] = "test-admin"
os.environ["DATABASE_URL"] = os.environ.get(
    "TEST_DATABASE_URL",
    "postgresql://telecoemprende:telecoemprende@localhost:5432/telecoemprende_test",
)

import app  # noqa: E402
import backend.services.admin as admin_service  # noqa: E402
import backend.services.registrations as registration_service  # noqa: E402
import backend.services.security as security_service  # noqa: E402
from backend.config import MAX_LOGIN_ATTEMPTS_PER_WINDOW  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


class ApiTestCase(unittest.TestCase):
    def setUp(self):
        admin_service.ADMIN_PASSWORD = "test-admin"
        security_service.request_log.clear()

        registration_service.init_db()

        conn = registration_service._get_connection()
        with conn.cursor() as cur:
            cur.execute("DELETE FROM registrations")
        conn.commit()
        conn.close()

        self.client = app.app.test_client()

    def register(self, email="juan@alumnos.upm.es"):
        return self.client.post(
            "/api/registrations",
            json={
                "nombre": "Juan",
                "apellidos": "Perez",
                "escuela": "ETSIT",
                "nivel": "Grado",
                "estudios": "Grado - Ingenieria de Tecnologias y Servicios de Telecomunicacion",
                "email": email,
                "telefono": "600123456",
                "departamento": "Tech/Ingeniería",
                "drive_link": "https://drive.google.com/file/d/test",
                "privacidad": True,
                "evento": "telecoemprende-2026-27",
                "telefono_oculto": "",
            },
        )

    def login(self, password="test-admin"):
        return self.client.post("/api/admin/login", json={"password": password})

    def test_registration_validation_error(self):
        response = self.client.post(
            "/api/registrations", json={"evento": "telecoemprende-2026-27"}
        )

        self.assertEqual(response.status_code, 400)
        payload = response.get_json()
        self.assertFalse(payload["ok"])
        self.assertIn("errors", payload)
        self.assertIn("email", payload["errors"])

    def test_duplicate_email_rejected_after_successful_registration(self):
        first = self.register()
        duplicate = self.register()

        self.assertEqual(first.status_code, 201)
        self.assertEqual(duplicate.status_code, 409)
        self.assertEqual(
            duplicate.get_json()["message"],
            "Ese correo ya está registrado.",
        )

    def test_admin_session_login_and_registrations_flow(self):
        self.register()

        session_before = self.client.get("/api/admin/session")
        login = self.login()
        session_after = self.client.get("/api/admin/session")
        registrations = self.client.get("/api/admin/registrations")

        self.assertEqual(session_before.status_code, 200)
        self.assertFalse(session_before.get_json()["authenticated"])
        self.assertEqual(login.status_code, 200)
        self.assertTrue(session_after.get_json()["authenticated"])
        self.assertEqual(registrations.status_code, 200)
        payload = registrations.get_json()
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["registros"][0]["email"], "juan@alumnos.upm.es")

    def test_admin_download_requires_auth_and_returns_file_when_authenticated(self):
        unauthorized = self.client.get("/api/admin/download")
        self.assertEqual(unauthorized.status_code, 401)

        self.register()
        self.login()
        authorized = self.client.get("/api/admin/download")

        self.assertEqual(authorized.status_code, 200)
        self.assertIn(
            "attachment; filename=registros_todos.xlsx",
            authorized.headers["Content-Disposition"],
        )
        authorized.close()

    def test_admin_login_rate_limited_after_repeated_failures(self):
        for _ in range(MAX_LOGIN_ATTEMPTS_PER_WINDOW):
            attempt = self.login(password="wrong-password")
            self.assertEqual(attempt.status_code, 401)

        blocked = self.login(password="wrong-password")
        self.assertEqual(blocked.status_code, 429)

        # Un intento correcto tras agotar el cupo también debe quedar bloqueado.
        still_blocked = self.login()
        self.assertEqual(still_blocked.status_code, 429)

    def test_admin_login_uses_its_own_rate_limit_bucket(self):
        # Agotar el cupo de /api/registrations no debe bloquear el login:
        # cada endpoint sensible tiene su propio contador por IP.
        from backend.config import MAX_REQUESTS_PER_MINUTE

        for _ in range(MAX_REQUESTS_PER_MINUTE):
            self.register()

        login = self.login()
        self.assertEqual(login.status_code, 200)

    def test_admin_update_rejects_invalid_departamento(self):
        self.register()
        self.login()
        registrations = self.client.get("/api/admin/registrations").get_json()
        reg_id = registrations["registros"][0]["id"]

        response = self.client.put(
            f"/api/admin/registrations/{reg_id}",
            json={
                "nombre": "Juan",
                "apellidos": "Perez",
                "escuela": "ETSIT",
                "nivel": "Grado",
                "estudios": "Grado - Test",
                "email": "juan@alumnos.upm.es",
                "telefono": "600123456",
                "departamento": "No Existe",
                "drive_link": "https://drive.google.com/file/d/test",
            },
        )

        self.assertEqual(response.status_code, 400)

    def test_admin_update_rejects_non_upm_email(self):
        self.register()
        self.login()
        registrations = self.client.get("/api/admin/registrations").get_json()
        reg_id = registrations["registros"][0]["id"]

        response = self.client.put(
            f"/api/admin/registrations/{reg_id}",
            json={
                "nombre": "Juan",
                "apellidos": "Perez",
                "escuela": "ETSIT",
                "nivel": "Grado",
                "estudios": "Grado - Test",
                "email": "juan@gmail.com",
                "telefono": "600123456",
                "departamento": "Tech/Ingeniería",
                "drive_link": "https://drive.google.com/file/d/test",
            },
        )

        self.assertEqual(response.status_code, 400)

    def test_admin_update_rejects_invalid_telefono(self):
        self.register()
        self.login()
        registrations = self.client.get("/api/admin/registrations").get_json()
        reg_id = registrations["registros"][0]["id"]

        response = self.client.put(
            f"/api/admin/registrations/{reg_id}",
            json={
                "nombre": "Juan",
                "apellidos": "Perez",
                "escuela": "ETSIT",
                "nivel": "Grado",
                "estudios": "Grado - Test",
                "email": "juan@alumnos.upm.es",
                "telefono": "no-es-un-telefono",
                "departamento": "Tech/Ingeniería",
                "drive_link": "https://drive.google.com/file/d/test",
            },
        )

        self.assertEqual(response.status_code, 400)

    def test_admin_update_estado_flow(self):
        self.register()
        self.login()
        registrations = self.client.get("/api/admin/registrations").get_json()
        reg = registrations["registros"][0]
        self.assertEqual(reg["estado"], "pendiente")
        self.assertFalse(reg["notificado"])

        response = self.client.patch(
            f"/api/admin/registrations/{reg['id']}/estado",
            json={"estado": "aceptado"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["ok"])

        registrations = self.client.get("/api/admin/registrations").get_json()
        self.assertEqual(registrations["registros"][0]["estado"], "aceptado")
        # Clasificar no envía el email por sí solo: hace falta el botón "Enviar".
        self.assertFalse(registrations["registros"][0]["notificado"])

    def test_admin_notificar_sends_only_pending_and_marks_notificado(self):
        self.register()
        self.login()
        reg_id = self.client.get("/api/admin/registrations").get_json()["registros"][0]["id"]
        self.client.patch(f"/api/admin/registrations/{reg_id}/estado", json={"estado": "aceptado"})

        response = self.client.post("/api/admin/registrations/notificar", json={"estado": "aceptado"})
        self.assertEqual(response.status_code, 200)
        payload = response.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["enviados"], 0)  # sin RESEND_API_KEY en tests, Resend no llega a llamarse

        # Sin RESEND_API_KEY el envío falla y notificado sigue en False (para poder reintentar).
        registrations = self.client.get("/api/admin/registrations").get_json()
        self.assertFalse(registrations["registros"][0]["notificado"])

        # Un segundo intento vuelve a considerarlo pendiente (no se marcó como enviado).
        response = self.client.post("/api/admin/registrations/notificar", json={"estado": "aceptado"})
        self.assertEqual(response.get_json()["total"], 1)

    def test_admin_notificar_rejects_pendiente(self):
        self.register()
        self.login()
        response = self.client.post("/api/admin/registrations/notificar", json={"estado": "pendiente"})
        self.assertEqual(response.status_code, 400)

    def test_admin_notificar_requires_auth(self):
        self.register()
        response = self.client.post("/api/admin/registrations/notificar", json={"estado": "aceptado"})
        self.assertEqual(response.status_code, 401)

    def test_admin_update_estado_rejects_invalid_value(self):
        self.register()
        self.login()
        registrations = self.client.get("/api/admin/registrations").get_json()
        reg_id = registrations["registros"][0]["id"]

        response = self.client.patch(
            f"/api/admin/registrations/{reg_id}/estado",
            json={"estado": "no-es-un-estado"},
        )
        self.assertEqual(response.status_code, 400)

    def test_admin_update_estado_requires_auth(self):
        self.register()
        self.login()
        registrations = self.client.get("/api/admin/registrations").get_json()
        reg_id = registrations["registros"][0]["id"]
        self.client.post("/api/admin/logout")

        response = self.client.patch(
            f"/api/admin/registrations/{reg_id}/estado",
            json={"estado": "aceptado"},
        )
        self.assertEqual(response.status_code, 401)

    def test_legacy_get_admin_logout_route_removed(self):
        # Era alcanzable con una navegación GET de nivel superior (cross-site),
        # forzando el logout del admin sin su intención. El logout real vive
        # únicamente en POST /api/admin/logout.
        response = self.client.get("/admin/logout")
        self.assertEqual(response.status_code, 404)

    def test_admin_update_requires_auth(self):
        self.register()
        self.login()
        registrations_as_admin = self.client.get("/api/admin/registrations").get_json()
        reg_id = registrations_as_admin["registros"][0]["id"]
        self.client.post("/api/admin/logout")

        response = self.client.put(
            f"/api/admin/registrations/{reg_id}",
            json={
                "nombre": "Hackeado",
                "apellidos": "Perez",
                "escuela": "ETSIT",
                "nivel": "Grado",
                "estudios": "Grado - Test",
                "email": "juan@alumnos.upm.es",
                "telefono": "600123456",
                "departamento": "Tech/Ingeniería",
                "drive_link": "https://drive.google.com/file/d/test",
            },
        )

        self.assertEqual(response.status_code, 401)

    def test_excel_export_neutralizes_formula_injection(self):
        self.client.post(
            "/api/registrations",
            json={
                "nombre": "=cmd|'/c calc'!A1",
                "apellidos": "Perez",
                "escuela": "ETSIT",
                "nivel": "Grado",
                "estudios": "Grado - Test",
                "email": "formula@alumnos.upm.es",
                "telefono": "600123456",
                "departamento": "Tech/Ingeniería",
                "drive_link": "https://drive.google.com/file/d/formula",
                "privacidad": True,
                "evento": "telecoemprende-2026-27",
                "telefono_oculto": "",
            },
        )
        self.login()

        excel_bytes = registration_service.generar_excel_en_memoria()
        wb = load_workbook(excel_bytes)
        ws = wb.active
        nombre_cell = ws.cell(row=2, column=1).value

        self.assertFalse(nombre_cell.startswith("="))
        self.assertTrue(nombre_cell.startswith("'"))

    def test_obtener_ip_real_ignores_spoofed_forwarded_header(self):
        with app.app.test_request_context(
            headers={"X-Forwarded-For": "=cmd|not-an-ip, 203.0.113.5"}
        ):
            self.assertEqual(security_service.obtener_ip_real(), "203.0.113.5")

        with app.app.test_request_context(
            headers={"X-Forwarded-For": "not-an-ip-either"}
        ):
            # Sin ningún valor válido en la cadena, cae al remote_addr real.
            self.assertNotEqual(
                security_service.obtener_ip_real(), "not-an-ip-either"
            )


if __name__ == "__main__":
    unittest.main()
